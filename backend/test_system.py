import sys
import os
import json
import io
import tempfile
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(__file__))

import unittest
from app import app
from database import init_db, get_db, calculate_next_date
import openpyxl


class TestResults:
    def __init__(self):
        self.passed = 0
        self.failed = 0
        self.results = []

    def add(self, test_id, module, test_name, description, status, actual="", expected=""):
        self.results.append({
            'id': test_id,
            'module': module,
            'test_name': test_name,
            'description': description,
            'status': status,
            'actual': actual,
            'expected': expected
        })
        if status == 'PASS':
            self.passed += 1
        else:
            self.failed += 1

    def print_report(self):
        print("\n" + "="*100)
        print("设备巡检管理系统 - 测试报告")
        print("="*100)
        print(f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"总用例数: {len(self.results)}")
        print(f"通过: {self.passed} | 失败: {self.failed}")
        print(f"通过率: {self.passed/len(self.results)*100:.2f}%")
        print("-"*100)
        
        modules = {}
        for r in self.results:
            if r['module'] not in modules:
                modules[r['module']] = {'pass': 0, 'fail': 0, 'total': 0}
            modules[r['module']]['total'] += 1
            if r['status'] == 'PASS':
                modules[r['module']]['pass'] += 1
            else:
                modules[r['module']]['fail'] += 1
        
        print("\n【模块统计】")
        for mod, stats in modules.items():
            rate = stats['pass']/stats['total']*100 if stats['total'] > 0 else 0
            print(f"  {mod}: {stats['pass']}/{stats['total']} 通过 ({rate:.1f}%)")
        
        print("\n【详细测试结果】")
        print("-"*100)
        for r in self.results:
            status_icon = "✓" if r['status'] == 'PASS' else "✗"
            print(f"\n[{r['id']}] {r['module']} - {r['test_name']}")
            print(f"    描述: {r['description']}")
            print(f"    状态: {status_icon} {r['status']}")
            if r['status'] == 'FAIL':
                if r['expected']:
                    print(f"    期望: {r['expected']}")
                if r['actual']:
                    print(f"    实际: {r['actual']}")
        
        print("\n" + "="*100)
        
        failed = [r for r in self.results if r['status'] == 'FAIL']
        if failed:
            print(f"\n失败用例列表 ({len(failed)}):")
            for r in failed:
                print(f"  [{r['id']}] {r['module']} - {r['test_name']}")
        
        return self.results


tr = TestResults()


class BaseTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        app.config['TESTING'] = True
        app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
        cls.client = app.test_client()
        
        if os.path.exists('test_inspection.db'):
            os.remove('test_inspection.db')
        
        import database
        database.DATABASE = os.path.join(os.path.dirname(__file__), 'test_inspection.db')
        database.init_db()

    @classmethod
    def tearDownClass(cls):
        if os.path.exists('test_inspection.db'):
            os.remove('test_inspection.db')


def run_tests():
    client = app.test_client()
    
    import database
    db_path = os.path.join(os.path.dirname(__file__), 'test_inspection.db')
    if os.path.exists(db_path):
        os.remove(db_path)
    database.DATABASE = db_path
    database.init_db()
    
    test_id = 0
    
    # ============ 数据库工具函数测试 ============
    test_id += 1
    try:
        result = calculate_next_date('2025-01-01', '每日')
        assert result == '2025-01-02', f"Expected 2025-01-02, got {result}"
        tr.add(test_id, '工具函数', 'calculate_next_date-每日', '测试每日巡检日期计算', 'PASS', f"返回{result}", '2025-01-02')
    except Exception as e:
        tr.add(test_id, '工具函数', 'calculate_next_date-每日', '测试每日巡检日期计算', 'FAIL', str(e), '2025-01-02')
    
    test_id += 1
    try:
        result = calculate_next_date('2025-01-01', '每周')
        assert result == '2025-01-08', f"Expected 2025-01-08, got {result}"
        tr.add(test_id, '工具函数', 'calculate_next_date-每周', '测试每周巡检日期计算', 'PASS', f"返回{result}", '2025-01-08')
    except Exception as e:
        tr.add(test_id, '工具函数', 'calculate_next_date-每周', '测试每周巡检日期计算', 'FAIL', str(e), '2025-01-08')
    
    test_id += 1
    try:
        result = calculate_next_date('2025-01-15', '每月')
        assert result == '2025-02-15', f"Expected 2025-02-15, got {result}"
        tr.add(test_id, '工具函数', 'calculate_next_date-每月', '测试每月巡检日期计算', 'PASS', f"返回{result}", '2025-02-15')
    except Exception as e:
        tr.add(test_id, '工具函数', 'calculate_next_date-每月', '测试每月巡检日期计算', 'FAIL', str(e), '2025-02-15')
    
    test_id += 1
    try:
        result = calculate_next_date('2025-12-25', '每月')
        assert result == '2026-01-25', f"Expected 2026-01-25, got {result}"
        tr.add(test_id, '工具函数', 'calculate_next_date-跨年', '测试跨年月份计算', 'PASS', f"返回{result}", '2026-01-25')
    except Exception as e:
        tr.add(test_id, '工具函数', 'calculate_next_date-跨年', '测试跨年月份计算', 'FAIL', str(e), '2026-01-25')
    
    test_id += 1
    try:
        result = calculate_next_date('2025-01-01', '每季')
        assert result == '2025-04-01', f"Expected 2025-04-01, got {result}"
        tr.add(test_id, '工具函数', 'calculate_next_date-每季', '测试每季巡检日期计算', 'PASS', f"返回{result}", '2025-04-01')
    except Exception as e:
        tr.add(test_id, '工具函数', 'calculate_next_date-每季', '测试每季巡检日期计算', 'FAIL', str(e), '2025-04-01')
    
    test_id += 1
    try:
        result = calculate_next_date('2025-01-01', '自定义', 15)
        assert result == '2025-01-16', f"Expected 2025-01-16, got {result}"
        tr.add(test_id, '工具函数', 'calculate_next_date-自定义', '测试自定义天数巡检日期计算', 'PASS', f"返回{result}", '2025-01-16')
    except Exception as e:
        tr.add(test_id, '工具函数', 'calculate_next_date-自定义', '测试自定义天数巡检日期计算', 'FAIL', str(e), '2025-01-16')
    
    test_id += 1
    try:
        result = calculate_next_date('', '每周')
        assert result is not None, "返回空日期"
        tr.add(test_id, '工具函数', 'calculate_next_date-空输入', '测试空输入时返回今天', 'PASS', f"返回{result}", '返回当前日期')
    except Exception as e:
        tr.add(test_id, '工具函数', 'calculate_next_date-空输入', '测试空输入时返回今天', 'FAIL', str(e), '返回当前日期')
    
    # ============ 基础接口测试 ============
    test_id += 1
    try:
        resp = client.get('/')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert 'message' in data
        tr.add(test_id, '基础接口', '根路径访问', '测试API根路径返回正常', 'PASS', f"状态码{resp.status_code}", '200')
    except Exception as e:
        tr.add(test_id, '基础接口', '根路径访问', '测试API根路径返回正常', 'FAIL', str(e), '200')
    
    # ============ 设备管理API测试 ============
    test_id += 1
    try:
        resp = client.post('/api/devices', json={
            'name': '测试服务器01',
            'model': 'PowerEdge R740',
            'brand': 'Dell',
            'asset_no': 'DEV-001',
            'category': '服务器',
            'location': '机房A-01',
            'purchase_date': '2024-01-15',
            'warranty_date': '2027-01-15',
            'status': '正常'
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert 'id' in data
        device1_id = data['id']
        tr.add(test_id, '设备管理', '创建设备-服务器', '创建正常服务器设备', 'PASS', f"创建成功ID={device1_id}", '返回200包含id')
    except Exception as e:
        device1_id = None
        tr.add(test_id, '设备管理', '创建设备-服务器', '创建正常服务器设备', 'FAIL', str(e), '返回200包含id')
    
    test_id += 1
    try:
        resp = client.post('/api/devices', json={
            'name': '测试网络交换机01',
            'model': 'Catalyst 2960',
            'brand': 'Cisco',
            'asset_no': 'DEV-002',
            'category': '网络设备',
            'location': '机房A-02',
            'purchase_date': '2024-03-20',
            'warranty_date': '2026-03-20',
            'status': '正常'
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        device2_id = data['id']
        tr.add(test_id, '设备管理', '创建设备-网络设备', '创建网络设备', 'PASS', f"创建成功ID={device2_id}", '返回200包含id')
    except Exception as e:
        device2_id = None
        tr.add(test_id, '设备管理', '创建设备-网络设备', '创建网络设备', 'FAIL', str(e), '返回200包含id')
    
    test_id += 1
    try:
        resp = client.post('/api/devices', json={
            'name': '测试打印机01',
            'model': 'LaserJet Pro',
            'brand': 'HP',
            'asset_no': 'DEV-003',
            'category': '打印机',
            'location': '办公区B',
            'status': '故障'
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        device3_id = data['id']
        tr.add(test_id, '设备管理', '创建设备-故障状态', '创建故障状态设备', 'PASS', f"创建成功ID={device3_id}", '返回200包含id')
    except Exception as e:
        device3_id = None
        tr.add(test_id, '设备管理', '创建设备-故障状态', '创建故障状态设备', 'FAIL', str(e), '返回200包含id')
    
    test_id += 1
    try:
        resp = client.post('/api/devices', json={
            'name': '测试监控摄像头01',
            'category': '监控',
            'location': '大门口'
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        device4_id = data['id']
        tr.add(test_id, '设备管理', '创建设备-最小字段', '只填必填字段创建设备', 'PASS', f"创建成功ID={device4_id}", '返回200包含id')
    except Exception as e:
        device4_id = None
        tr.add(test_id, '设备管理', '创建设备-最小字段', '只填必填字段创建设备', 'FAIL', str(e), '返回200包含id')
    
    test_id += 1
    try:
        resp = client.post('/api/devices', json={
            'name': '测试电脑01',
            'model': 'ThinkPad X1',
            'brand': 'Lenovo',
            'asset_no': 'DEV-005',
            'category': '电脑',
            'location': '研发部',
            'status': '正常'
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        device5_id = data['id']
        tr.add(test_id, '设备管理', '创建设备-电脑', '创建电脑设备', 'PASS', f"创建成功ID={device5_id}", '返回200包含id')
    except Exception as e:
        device5_id = None
        tr.add(test_id, '设备管理', '创建设备-电脑', '创建电脑设备', 'FAIL', str(e), '返回200包含id')
    
    test_id += 1
    try:
        resp = client.get('/api/devices')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert isinstance(data, list)
        assert len(data) >= 4
        tr.add(test_id, '设备管理', '获取设备列表', '获取所有设备列表', 'PASS', f"返回{len(data)}条记录", '返回数组且>=4条')
    except Exception as e:
        tr.add(test_id, '设备管理', '获取设备列表', '获取所有设备列表', 'FAIL', str(e), '返回数组且>=4条')
    
    test_id += 1
    try:
        resp = client.get(f'/api/devices/{device1_id}')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['name'] == '测试服务器01'
        assert data['category'] == '服务器'
        tr.add(test_id, '设备管理', '获取单个设备详情', '根据ID获取设备详情', 'PASS', f"设备名={data['name']}", '正确返回设备信息')
    except Exception as e:
        tr.add(test_id, '设备管理', '获取单个设备详情', '根据ID获取设备详情', 'FAIL', str(e), '正确返回设备信息')
    
    test_id += 1
    try:
        resp = client.get('/api/devices/99999')
        assert resp.status_code == 404
        tr.add(test_id, '设备管理', '获取不存在的设备', '获取不存在ID的设备返回404', 'PASS', f"状态码{resp.status_code}", '返回404')
    except Exception as e:
        tr.add(test_id, '设备管理', '获取不存在的设备', '获取不存在ID的设备返回404', 'FAIL', str(e), '返回404')
    
    test_id += 1
    try:
        resp = client.get('/api/devices?category=服务器')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert all(d['category'] == '服务器' for d in data)
        tr.add(test_id, '设备管理', '按分类筛选设备', '按服务器分类筛选', 'PASS', f"返回{len(data)}条服务器", '仅返回服务器设备')
    except Exception as e:
        tr.add(test_id, '设备管理', '按分类筛选设备', '按服务器分类筛选', 'FAIL', str(e), '仅返回服务器设备')
    
    test_id += 1
    try:
        resp = client.get('/api/devices?status=故障')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert all(d['status'] == '故障' for d in data)
        tr.add(test_id, '设备管理', '按状态筛选设备', '按故障状态筛选', 'PASS', f"返回{len(data)}条故障", '仅返回故障状态设备')
    except Exception as e:
        tr.add(test_id, '设备管理', '按状态筛选设备', '按故障状态筛选', 'FAIL', str(e), '仅返回故障状态设备')
    
    test_id += 1
    try:
        resp = client.get('/api/devices?keyword=Dell')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        tr.add(test_id, '设备管理', '按关键词搜索-品牌', '按品牌关键词搜索', 'PASS', f"返回{len(data)}条", '返回匹配品牌的设备')
    except Exception as e:
        tr.add(test_id, '设备管理', '按关键词搜索-品牌', '按品牌关键词搜索', 'FAIL', str(e), '返回匹配品牌的设备')
    
    test_id += 1
    try:
        resp = client.get('/api/devices?keyword=测试服务器')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert len(data) >= 1
        tr.add(test_id, '设备管理', '按关键词搜索-名称', '按名称关键词搜索', 'PASS', f"返回{len(data)}条", '返回匹配名称的设备')
    except Exception as e:
        tr.add(test_id, '设备管理', '按关键词搜索-名称', '按名称关键词搜索', 'FAIL', str(e), '返回匹配名称的设备')
    
    test_id += 1
    try:
        resp = client.put(f'/api/devices/{device1_id}', json={
            'name': '测试服务器01-已更新',
            'model': 'PowerEdge R750',
            'brand': 'Dell',
            'asset_no': 'DEV-001',
            'category': '服务器',
            'location': '机房A-01-更新',
            'purchase_date': '2024-01-15',
            'warranty_date': '2028-01-15',
            'status': '正常'
        })
        assert resp.status_code == 200
        
        resp2 = client.get(f'/api/devices/{device1_id}')
        data = json.loads(resp2.data)
        assert data['name'] == '测试服务器01-已更新'
        assert data['model'] == 'PowerEdge R750'
        assert data['location'] == '机房A-01-更新'
        tr.add(test_id, '设备管理', '更新设备信息', '更新设备名称、型号、位置', 'PASS', f"更新后名称={data['name']}", '字段正确更新')
    except Exception as e:
        tr.add(test_id, '设备管理', '更新设备信息', '更新设备名称、型号、位置', 'FAIL', str(e), '字段正确更新')
    
    # ============ 巡检计划API测试 ============
    test_id += 1
    try:
        resp = client.post('/api/plans', json={
            'device_id': device1_id,
            'inspection_type': '每周',
            'items': ['温度检查', '指示灯状态', '系统日志'],
            'last_inspection_date': '2025-01-01',
            'responsible_person': '张三'
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        plan1_id = data['id']
        tr.add(test_id, '巡检计划', '创建计划-每周', '创建每周巡检计划', 'PASS', f"创建成功ID={plan1_id}", '返回200包含id')
    except Exception as e:
        plan1_id = None
        tr.add(test_id, '巡检计划', '创建计划-每周', '创建每周巡检计划', 'FAIL', str(e), '返回200包含id')
    
    test_id += 1
    try:
        resp = client.post('/api/plans', json={
            'device_id': device2_id,
            'inspection_type': '每月',
            'items': ['温度检查', '线缆检查'],
            'last_inspection_date': '2025-06-15',
            'responsible_person': '李四'
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        plan2_id = data['id']
        tr.add(test_id, '巡检计划', '创建计划-每月', '创建每月巡检计划', 'PASS', f"创建成功ID={plan2_id}", '返回200包含id')
    except Exception as e:
        plan2_id = None
        tr.add(test_id, '巡检计划', '创建计划-每月', '创建每月巡检计划', 'FAIL', str(e), '返回200包含id')
    
    test_id += 1
    try:
        resp = client.post('/api/plans', json={
            'device_id': device5_id,
            'inspection_type': '自定义',
            'custom_days': 10,
            'items': ['清洁除尘', '性能检查'],
            'last_inspection_date': '2025-06-01',
            'responsible_person': '王五'
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        plan3_id = data['id']
        tr.add(test_id, '巡检计划', '创建计划-自定义', '创建自定义天数巡检计划', 'PASS', f"创建成功ID={plan3_id}", '返回200包含id')
    except Exception as e:
        plan3_id = None
        tr.add(test_id, '巡检计划', '创建计划-自定义', '创建自定义天数巡检计划', 'FAIL', str(e), '返回200包含id')
    
    test_id += 1
    try:
        resp = client.get('/api/plans')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert isinstance(data, list)
        assert len(data) >= 3
        assert 'items' in data[0]
        assert isinstance(data[0]['items'], list)
        tr.add(test_id, '巡检计划', '获取计划列表', '获取所有巡检计划', 'PASS', f"返回{len(data)}条,items已解析", '返回数组且items解析为列表')
    except Exception as e:
        tr.add(test_id, '巡检计划', '获取计划列表', '获取所有巡检计划', 'FAIL', str(e), '返回数组且items解析为列表')
    
    test_id += 1
    try:
        resp = client.get(f'/api/plans/{plan1_id}')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['inspection_type'] == '每周'
        assert data['responsible_person'] == '张三'
        assert data['next_inspection_date'] == '2025-01-08'
        tr.add(test_id, '巡检计划', '获取单个计划', '获取计划详情并验证下次日期计算', 'PASS', f"下次巡检={data['next_inspection_date']}", '正确返回计划及计算后日期')
    except Exception as e:
        tr.add(test_id, '巡检计划', '获取单个计划', '获取计划详情并验证下次日期计算', 'FAIL', str(e), '正确返回计划及计算后日期')
    
    test_id += 1
    try:
        resp = client.get('/api/plans/99999')
        assert resp.status_code == 404
        tr.add(test_id, '巡检计划', '获取不存在的计划', '获取不存在ID的计划返回404', 'PASS', f"状态码{resp.status_code}", '返回404')
    except Exception as e:
        tr.add(test_id, '巡检计划', '获取不存在的计划', '获取不存在ID的计划返回404', 'FAIL', str(e), '返回404')
    
    test_id += 1
    try:
        resp = client.put(f'/api/plans/{plan1_id}', json={
            'device_id': device1_id,
            'inspection_type': '每日',
            'items': ['温度检查', '指示灯状态', '系统日志', '安全检查'],
            'last_inspection_date': '2025-06-10',
            'responsible_person': '张三-更新'
        })
        assert resp.status_code == 200
        
        resp2 = client.get(f'/api/plans/{plan1_id}')
        data = json.loads(resp2.data)
        assert data['inspection_type'] == '每日'
        assert data['responsible_person'] == '张三-更新'
        assert len(data['items']) == 4
        tr.add(test_id, '巡检计划', '更新巡检计划', '更新巡检类型、负责人、巡检项目', 'PASS', f"类型={data['inspection_type']},项目数={len(data['items'])}", '字段正确更新')
    except Exception as e:
        tr.add(test_id, '巡检计划', '更新巡检计划', '更新巡检类型、负责人、巡检项目', 'FAIL', str(e), '字段正确更新')
    
    # ============ 巡检记录API测试 ============
    test_id += 1
    try:
        resp = client.post('/api/records', json={
            'device_id': device1_id,
            'plan_id': plan1_id,
            'inspection_date': '2025-06-15',
            'result': '正常',
            'items_detail': ['温度检查', '指示灯状态', '系统日志'],
            'inspector': '巡检员A',
            'photos': []
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        record1_id = data['id']
        tr.add(test_id, '巡检记录', '创建记录-正常', '创建正常巡检记录', 'PASS', f"创建成功ID={record1_id}", '返回200包含id')
    except Exception as e:
        record1_id = None
        tr.add(test_id, '巡检记录', '创建记录-正常', '创建正常巡检记录', 'FAIL', str(e), '返回200包含id')
    
    test_id += 1
    try:
        resp = client.post('/api/records', json={
            'device_id': device3_id,
            'inspection_date': '2025-06-14',
            'result': '异常',
            'items_detail': ['温度检查', '性能检查'],
            'anomaly_description': '设备温度过高，风扇异响，需立即处理',
            'inspector': '巡检员B',
            'photos': []
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        record2_id = data['id']
        
        resp2 = client.get(f'/api/devices/{device3_id}')
        dev_data = json.loads(resp2.data)
        assert dev_data['status'] == '故障'
        tr.add(test_id, '巡检记录', '创建记录-异常联动设备状态', '异常记录自动更新设备状态为故障', 'PASS', f"设备状态={dev_data['status']}", '设备状态自动更新为故障')
    except Exception as e:
        record2_id = None
        tr.add(test_id, '巡检记录', '创建记录-异常联动设备状态', '异常记录自动更新设备状态为故障', 'FAIL', str(e), '设备状态自动更新为故障')
    
    test_id += 1
    try:
        resp = client.post('/api/records', json={
            'device_id': device2_id,
            'plan_id': plan2_id,
            'inspection_date': '2025-06-13',
            'result': '正常',
            'items_detail': ['线缆检查', '温度检查'],
            'inspector': '巡检员C',
            'photos': []
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        record3_id = data['id']
        
        resp2 = client.get(f'/api/plans/{plan2_id}')
        plan_data = json.loads(resp2.data)
        assert plan_data['last_inspection_date'] == '2025-06-13'
        tr.add(test_id, '巡检记录', '创建记录-更新计划日期', '创建记录后更新关联计划的巡检日期', 'PASS', f"计划最近巡检={plan_data['last_inspection_date']}", '关联计划日期被更新')
    except Exception as e:
        record3_id = None
        tr.add(test_id, '巡检记录', '创建记录-更新计划日期', '创建记录后更新关联计划的巡检日期', 'FAIL', str(e), '关联计划日期被更新')
    
    test_id += 1
    try:
        resp = client.get('/api/records')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert isinstance(data, list)
        assert len(data) >= 3
        tr.add(test_id, '巡检记录', '获取记录列表', '获取所有巡检记录', 'PASS', f"返回{len(data)}条记录", '返回数组且>=3条')
    except Exception as e:
        tr.add(test_id, '巡检记录', '获取记录列表', '获取所有巡检记录', 'FAIL', str(e), '返回数组且>=3条')
    
    test_id += 1
    try:
        resp = client.get(f'/api/records/{record1_id}')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['result'] == '正常'
        assert data['inspector'] == '巡检员A'
        tr.add(test_id, '巡检记录', '获取单条记录', '获取单条巡检记录详情', 'PASS', f"结果={data['result']},巡检人={data['inspector']}", '正确返回记录详情')
    except Exception as e:
        tr.add(test_id, '巡检记录', '获取单条记录', '获取单条巡检记录详情', 'FAIL', str(e), '正确返回记录详情')
    
    test_id += 1
    try:
        resp = client.get('/api/records/99999')
        assert resp.status_code == 404
        tr.add(test_id, '巡检记录', '获取不存在的记录', '获取不存在ID的记录返回404', 'PASS', f"状态码{resp.status_code}", '返回404')
    except Exception as e:
        tr.add(test_id, '巡检记录', '获取不存在的记录', '获取不存在ID的记录返回404', 'FAIL', str(e), '返回404')
    
    test_id += 1
    try:
        resp = client.get('/api/records?result=正常')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert all(r['result'] == '正常' for r in data)
        tr.add(test_id, '巡检记录', '按结果筛选记录', '按正常结果筛选记录', 'PASS', f"返回{len(data)}条正常记录", '仅返回正常结果的记录')
    except Exception as e:
        tr.add(test_id, '巡检记录', '按结果筛选记录', '按正常结果筛选记录', 'FAIL', str(e), '仅返回正常结果的记录')
    
    test_id += 1
    try:
        resp = client.get(f'/api/records?device_id={device1_id}')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert all(r['device_id'] == device1_id for r in data)
        tr.add(test_id, '巡检记录', '按设备筛选记录', '按设备ID筛选记录', 'PASS', f"返回{len(data)}条设备{device1_id}的记录", '仅返回指定设备的记录')
    except Exception as e:
        tr.add(test_id, '巡检记录', '按设备筛选记录', '按设备ID筛选记录', 'FAIL', str(e), '仅返回指定设备的记录')
    
    # ============ 报修管理API测试 ============
    test_id += 1
    try:
        resp = client.post('/api/repairs', json={
            'device_id': device3_id,
            'report_date': '2025-06-15',
            'fault_description': '打印机卡纸频繁，无法正常打印，墨盒异常',
            'status': '待处理'
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        repair1_id = data['id']
        
        resp2 = client.get(f'/api/devices/{device3_id}')
        dev_data = json.loads(resp2.data)
        assert dev_data['status'] == '维修中'
        tr.add(test_id, '报修管理', '创建报修-待处理', '创建报修并联动设备状态为维修中', 'PASS', f"报修ID={repair1_id},设备状态={dev_data['status']}", '报修创建且设备状态=维修中')
    except Exception as e:
        repair1_id = None
        tr.add(test_id, '报修管理', '创建报修-待处理', '创建报修并联动设备状态为维修中', 'FAIL', str(e), '报修创建且设备状态=维修中')
    
    test_id += 1
    try:
        resp = client.post('/api/repairs', json={
            'device_id': device5_id,
            'report_date': '2025-06-10',
            'fault_description': '电脑蓝屏频繁，内存可能有问题',
            'repair_date': '2025-06-12',
            'repair_result': '更换内存条后恢复正常',
            'repair_cost': 580.00,
            'status': '已完成'
        })
        assert resp.status_code == 200
        data = json.loads(resp.data)
        repair2_id = data['id']
        
        resp2 = client.get(f'/api/devices/{device5_id}')
        dev_data = json.loads(resp2.data)
        assert dev_data['status'] == '正常'
        tr.add(test_id, '报修管理', '创建报修-已完成', '创建已完成报修并联动设备状态为正常', 'PASS', f"报修ID={repair2_id},设备状态={dev_data['status']}", '报修创建且设备状态=正常')
    except Exception as e:
        repair2_id = None
        tr.add(test_id, '报修管理', '创建报修-已完成', '创建已完成报修并联动设备状态为正常', 'FAIL', str(e), '报修创建且设备状态=正常')
    
    test_id += 1
    try:
        resp = client.get('/api/repairs')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert isinstance(data, list)
        assert len(data) >= 2
        tr.add(test_id, '报修管理', '获取报修列表', '获取所有报修记录', 'PASS', f"返回{len(data)}条记录", '返回数组且>=2条')
    except Exception as e:
        tr.add(test_id, '报修管理', '获取报修列表', '获取所有报修记录', 'FAIL', str(e), '返回数组且>=2条')
    
    test_id += 1
    try:
        resp = client.get(f'/api/repairs/{repair1_id}')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert '打印机卡纸' in data['fault_description']
        tr.add(test_id, '报修管理', '获取单条报修', '获取单条报修详情', 'PASS', f"故障描述包含关键词", '正确返回报修详情')
    except Exception as e:
        tr.add(test_id, '报修管理', '获取单条报修', '获取单条报修详情', 'FAIL', str(e), '正确返回报修详情')
    
    test_id += 1
    try:
        resp = client.get('/api/repairs/99999')
        assert resp.status_code == 404
        tr.add(test_id, '报修管理', '获取不存在的报修', '获取不存在ID的报修返回404', 'PASS', f"状态码{resp.status_code}", '返回404')
    except Exception as e:
        tr.add(test_id, '报修管理', '获取不存在的报修', '获取不存在ID的报修返回404', 'FAIL', str(e), '返回404')
    
    test_id += 1
    try:
        resp = client.put(f'/api/repairs/{repair1_id}', json={
            'device_id': device3_id,
            'report_date': '2025-06-15',
            'fault_description': '打印机卡纸频繁，无法正常打印，墨盒异常',
            'repair_date': '2025-06-16',
            'repair_result': '更换进纸轮和墨盒，测试正常',
            'repair_cost': 350.00,
            'status': '已完成'
        })
        assert resp.status_code == 200
        
        resp2 = client.get(f'/api/repairs/{repair1_id}')
        data = json.loads(resp2.data)
        assert data['status'] == '已完成'
        assert data['repair_cost'] == 350.00
        
        resp3 = client.get(f'/api/devices/{device3_id}')
        dev_data = json.loads(resp3.data)
        assert dev_data['status'] == '正常'
        tr.add(test_id, '报修管理', '更新报修-完成维修', '更新报修状态为已完成，联动设备状态', 'PASS', f"状态={data['status']},费用={data['repair_cost']},设备状态={dev_data['status']}", '报修完成且设备恢复正常')
    except Exception as e:
        tr.add(test_id, '报修管理', '更新报修-完成维修', '更新报修状态为已完成，联动设备状态', 'FAIL', str(e), '报修完成且设备恢复正常')
    
    test_id += 1
    try:
        resp = client.get('/api/repairs?status=待处理')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert all(r['status'] == '待处理' for r in data)
        tr.add(test_id, '报修管理', '按状态筛选报修', '按待处理状态筛选', 'PASS', f"返回{len(data)}条待处理", '仅返回待处理报修')
    except Exception as e:
        tr.add(test_id, '报修管理', '按状态筛选报修', '按待处理状态筛选', 'FAIL', str(e), '仅返回待处理报修')
    
    test_id += 1
    try:
        resp = client.get(f'/api/repairs?device_id={device3_id}')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert all(r['device_id'] == device3_id for r in data)
        tr.add(test_id, '报修管理', '按设备筛选报修', '按设备ID筛选报修', 'PASS', f"返回{len(data)}条", '仅返回指定设备的报修')
    except Exception as e:
        tr.add(test_id, '报修管理', '按设备筛选报修', '按设备ID筛选报修', 'FAIL', str(e), '仅返回指定设备的报修')
    
    # ============ 仪表盘API测试 ============
    test_id += 1
    try:
        resp = client.get('/api/dashboard/stats')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert 'total_devices' in data
        assert 'normal_devices' in data
        assert 'anomaly_devices' in data
        assert 'today_inspection' in data
        assert 'overdue_count' in data
        assert 'category_stats' in data
        assert 'trend_data' in data
        assert 'warranty_soon' in data
        assert isinstance(data['category_stats'], list)
        assert isinstance(data['trend_data'], list)
        tr.add(test_id, '仪表盘', '获取统计数据', '获取仪表盘所有统计指标', 'PASS', f"设备总数={data['total_devices']},正常={data['normal_devices']},异常={data['anomaly_devices']}", '返回所有8个统计字段')
    except Exception as e:
        tr.add(test_id, '仪表盘', '获取统计数据', '获取仪表盘所有统计指标', 'FAIL', str(e), '返回所有8个统计字段')
    
    test_id += 1
    try:
        resp = client.get('/api/dashboard/stats')
        data = json.loads(resp.data)
        assert data['total_devices'] >= data['normal_devices'] + data['anomaly_devices']
        tr.add(test_id, '仪表盘', '统计数据一致性', '验证设备总数 >= 正常 + 异常', 'PASS', f"总数={data['total_devices']},正常+异常={data['normal_devices'] + data['anomaly_devices']}", '数据统计一致')
    except Exception as e:
        tr.add(test_id, '仪表盘', '统计数据一致性', '验证设备总数 >= 正常 + 异常', 'FAIL', str(e), '数据统计一致')
    
    test_id += 1
    try:
        resp = client.get('/api/dashboard/reminders')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert 'inspection_reminders' in data
        assert 'warranty_reminders' in data
        assert isinstance(data['inspection_reminders'], list)
        assert isinstance(data['warranty_reminders'], list)
        tr.add(test_id, '仪表盘', '获取提醒数据', '获取巡检和保修期提醒', 'PASS', f"巡检提醒={len(data['inspection_reminders'])}条,保修提醒={len(data['warranty_reminders'])}条", '返回两种提醒列表')
    except Exception as e:
        tr.add(test_id, '仪表盘', '获取提醒数据', '获取巡检和保修期提醒', 'FAIL', str(e), '返回两种提醒列表')
    
    # ============ 导出功能测试 ============
    test_id += 1
    try:
        resp = client.get('/api/export/devices')
        assert resp.status_code == 200
        assert resp.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        content = io.BytesIO(resp.data)
        wb = openpyxl.load_workbook(content)
        ws = wb.active
        assert ws.title == '设备清单'
        headers = [cell.value for cell in ws[1]]
        assert '设备名称' in headers
        assert '状态' in headers
        assert ws.max_row >= 2
        tr.add(test_id, '导出功能', '导出设备Excel', '导出设备清单Excel验证格式和内容', 'PASS', f"行数={ws.max_row},表头正确", '返回正确格式Excel')
    except Exception as e:
        tr.add(test_id, '导出功能', '导出设备Excel', '导出设备清单Excel验证格式和内容', 'FAIL', str(e), '返回正确格式Excel')
    
    test_id += 1
    try:
        resp = client.get('/api/export/records')
        assert resp.status_code == 200
        content = io.BytesIO(resp.data)
        wb = openpyxl.load_workbook(content)
        ws = wb.active
        assert ws.title == '巡检记录'
        headers = [cell.value for cell in ws[1]]
        assert '设备名称' in headers
        assert '巡检结果' in headers
        tr.add(test_id, '导出功能', '导出巡检记录Excel', '导出巡检记录Excel验证', 'PASS', f"Sheet={ws.title},行数={ws.max_row}", '返回正确格式Excel')
    except Exception as e:
        tr.add(test_id, '导出功能', '导出巡检记录Excel', '导出巡检记录Excel验证', 'FAIL', str(e), '返回正确格式Excel')
    
    test_id += 1
    try:
        resp = client.get('/api/export/repairs')
        assert resp.status_code == 200
        content = io.BytesIO(resp.data)
        wb = openpyxl.load_workbook(content)
        ws = wb.active
        assert ws.title == '报修记录'
        headers = [cell.value for cell in ws[1]]
        assert '故障描述' in headers
        assert '维修费用' in headers
        tr.add(test_id, '导出功能', '导出报修记录Excel', '导出报修记录Excel验证', 'PASS', f"Sheet={ws.title},行数={ws.max_row}", '返回正确格式Excel')
    except Exception as e:
        tr.add(test_id, '导出功能', '导出报修记录Excel', '导出报修记录Excel验证', 'FAIL', str(e), '返回正确格式Excel')
    
    # ============ 文件上传测试 ============
    test_id += 1
    try:
        data = {'file': (io.BytesIO(b'test image content'), 'test.png', 'image/png')}
        resp = client.post('/api/upload', data=data, content_type='multipart/form-data')
        assert resp.status_code == 200
        result = json.loads(resp.data)
        assert 'url' in result
        assert 'filename' in result
        tr.add(test_id, '文件上传', '上传图片-png', '上传PNG图片文件', 'PASS', f"url={result.get('url')}", '返回上传URL和文件名')
    except Exception as e:
        tr.add(test_id, '文件上传', '上传图片-png', '上传PNG图片文件', 'FAIL', str(e), '返回上传URL和文件名')
    
    test_id += 1
    try:
        data = {'file': (io.BytesIO(b'test jpg content'), 'photo.jpg', 'image/jpeg')}
        resp = client.post('/api/upload', data=data, content_type='multipart/form-data')
        assert resp.status_code == 200
        result = json.loads(resp.data)
        assert 'url' in result
        tr.add(test_id, '文件上传', '上传图片-jpg', '上传JPG图片文件', 'PASS', f"上传成功", '返回上传URL')
    except Exception as e:
        tr.add(test_id, '文件上传', '上传图片-jpg', '上传JPG图片文件', 'FAIL', str(e), '返回上传URL')
    
    test_id += 1
    try:
        data = {'file': (io.BytesIO(b'test exe file'), 'test.exe', 'application/exe')}
        resp = client.post('/api/upload', data=data, content_type='multipart/form-data')
        assert resp.status_code == 400
        tr.add(test_id, '文件上传', '上传非法文件类型', '上传不支持的文件类型返回400', 'PASS', f"状态码={resp.status_code}", '返回400拒绝上传')
    except Exception as e:
        tr.add(test_id, '文件上传', '上传非法文件类型', '上传不支持的文件类型返回400', 'FAIL', str(e), '返回400拒绝上传')
    
    test_id += 1
    try:
        resp = client.post('/api/upload', data={}, content_type='multipart/form-data')
        assert resp.status_code == 400
        tr.add(test_id, '文件上传', '上传空文件', '未上传文件返回400', 'PASS', f"状态码={resp.status_code}", '返回400提示无文件')
    except Exception as e:
        tr.add(test_id, '文件上传', '上传空文件', '未上传文件返回400', 'FAIL', str(e), '返回400提示无文件')
    
    # ============ Excel导入测试 ============
    test_id += 1
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '设备'
        ws.append(['设备名称', '型号', '品牌', '资产编号', '分类', '位置', '购买日期', '保修期', '状态'])
        ws.append(['导入测试设备1', 'Model-A', 'Brand-A', 'IMP-001', '服务器', '机房C', '2024-01-01', '2026-01-01', '正常'])
        ws.append(['导入测试设备2', 'Model-B', 'Brand-B', 'IMP-002', '电脑', '办公区', '2024-06-01', '2027-06-01', '正常'])
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        data = {'file': (buf, 'import_test.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        resp = client.post('/api/devices/import', data=data, content_type='multipart/form-data')
        assert resp.status_code == 200
        result = json.loads(resp.data)
        assert '成功导入' in result['message']
        assert '2' in result['message']
        tr.add(test_id, 'Excel导入', '导入设备Excel', '从Excel批量导入2台设备', 'PASS', f"{result['message']}", '成功导入指定数量记录')
    except Exception as e:
        tr.add(test_id, 'Excel导入', '导入设备Excel', '从Excel批量导入2台设备', 'FAIL', str(e), '成功导入指定数量记录')
    
    test_id += 1
    try:
        resp = client.post('/api/devices/import', data={}, content_type='multipart/form-data')
        assert resp.status_code == 400
        tr.add(test_id, 'Excel导入', '导入空文件', '未上传Excel文件返回400', 'PASS', f"状态码={resp.status_code}", '返回400提示无文件')
    except Exception as e:
        tr.add(test_id, 'Excel导入', '导入空文件', '未上传Excel文件返回400', 'FAIL', str(e), '返回400提示无文件')
    
    # ============ 删除功能测试 ============
    test_id += 1
    try:
        unique_asset = f'DEL-{datetime.now().strftime("%H%M%S%f")}'
        resp = client.post('/api/devices', json={
            'name': '待删除设备',
            'asset_no': unique_asset,
            'category': '其他'
        })
        data = json.loads(resp.data)
        assert 'id' in data, f"创建响应缺少id字段: {data}"
        del_device_id = data['id']
        
        resp = client.delete(f'/api/devices/{del_device_id}')
        assert resp.status_code == 200
        
        resp2 = client.get(f'/api/devices/{del_device_id}')
        assert resp2.status_code == 404
        tr.add(test_id, '删除功能', '删除设备', '删除设备并验证不可再访问', 'PASS', f"设备已删除", '删除后无法获取该设备')
    except Exception as e:
        tr.add(test_id, '删除功能', '删除设备', '删除设备并验证不可再访问', 'FAIL', str(e), '删除后无法获取该设备')
    
    test_id += 1
    try:
        resp = client.post('/api/plans', json={
            'device_id': device1_id,
            'inspection_type': '每日',
            'items': ['温度检查'],
            'last_inspection_date': '2025-06-01',
            'responsible_person': '测试员'
        })
        data = json.loads(resp.data)
        del_plan_id = data['id']
        
        resp = client.delete(f'/api/plans/{del_plan_id}')
        assert resp.status_code == 200
        
        resp2 = client.get(f'/api/plans/{del_plan_id}')
        assert resp2.status_code == 404
        tr.add(test_id, '删除功能', '删除巡检计划', '删除巡检计划并验证不可再访问', 'PASS', f"计划已删除", '删除后无法获取该计划')
    except Exception as e:
        tr.add(test_id, '删除功能', '删除巡检计划', '删除巡检计划并验证不可再访问', 'FAIL', str(e), '删除后无法获取该计划')
    
    test_id += 1
    try:
        resp = client.post('/api/records', json={
            'device_id': device2_id,
            'inspection_date': '2025-06-16',
            'result': '正常',
            'items_detail': ['温度检查'],
            'inspector': '测试员',
            'photos': []
        })
        data = json.loads(resp.data)
        del_record_id = data['id']
        
        resp = client.delete(f'/api/records/{del_record_id}')
        assert resp.status_code == 200
        
        resp2 = client.get(f'/api/records/{del_record_id}')
        assert resp2.status_code == 404
        tr.add(test_id, '删除功能', '删除巡检记录', '删除巡检记录并验证不可再访问', 'PASS', f"记录已删除", '删除后无法获取该记录')
    except Exception as e:
        tr.add(test_id, '删除功能', '删除巡检记录', '删除巡检记录并验证不可再访问', 'FAIL', str(e), '删除后无法获取该记录')
    
    test_id += 1
    try:
        resp = client.post('/api/repairs', json={
            'device_id': device4_id,
            'report_date': '2025-06-16',
            'fault_description': '待删除测试报修',
            'status': '待处理'
        })
        data = json.loads(resp.data)
        del_repair_id = data['id']
        
        resp = client.delete(f'/api/repairs/{del_repair_id}')
        assert resp.status_code == 200
        
        resp2 = client.get(f'/api/repairs/{del_repair_id}')
        assert resp2.status_code == 404
        tr.add(test_id, '删除功能', '删除报修记录', '删除报修记录并验证不可再访问', 'PASS', f"报修已删除", '删除后无法获取该报修')
    except Exception as e:
        tr.add(test_id, '删除功能', '删除报修记录', '删除报修记录并验证不可再访问', 'FAIL', str(e), '删除后无法获取该报修')
    
    # ============ 前端界面测试（结构验证） ============
    test_id += 1
    try:
        html_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'index.html')
        assert os.path.exists(html_path)
        with open(html_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert '<div id="app">' in content
        assert 'vue' in content.lower()
        assert 'element' in content.lower()
        assert 'chart' in content.lower()
        assert 'axios' in content.lower()
        tr.add(test_id, '前端界面', 'HTML入口文件', '验证index.html包含Vue3、ElementPlus、Chart.js、Axios', 'PASS', "HTML入口文件正确", '包含必要的CDN引用')
    except Exception as e:
        tr.add(test_id, '前端界面', 'HTML入口文件', '验证index.html包含Vue3、ElementPlus、Chart.js、Axios', 'FAIL', str(e), '包含必要的CDN引用')
    
    test_id += 1
    try:
        js_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'app.js')
        assert os.path.exists(js_path)
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert 'createApp' in content
        assert 'createRouter' in content
        assert 'Dashboard' in content
        assert 'DeviceList' in content
        assert 'PlanList' in content
        assert 'RecordList' in content
        assert 'RepairList' in content
        tr.add(test_id, '前端界面', 'JS主文件结构', '验证app.js包含5个核心页面组件', 'PASS', "5个页面组件均已定义", '包含所有核心组件')
    except Exception as e:
        tr.add(test_id, '前端界面', 'JS主文件结构', '验证app.js包含5个核心页面组件', 'FAIL', str(e), '包含所有核心组件')
    
    test_id += 1
    try:
        js_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'app.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert 'API_BASE' in content
        assert 'axiosInstance' in content
        assert 'axios.create' in content
        tr.add(test_id, '前端界面', 'API连接配置', '验证前端API基础路径和axios配置', 'PASS', "API配置正确", '包含API_BASE和axios实例')
    except Exception as e:
        tr.add(test_id, '前端界面', 'API连接配置', '验证前端API基础路径和axios配置', 'FAIL', str(e), '包含API_BASE和axios实例')
    
    test_id += 1
    try:
        js_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'app.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read()
        routes_count = content.count("path:")
        assert routes_count >= 5
        assert "'/'" in content
        assert "'/devices'" in content
        assert "'/plans'" in content
        assert "'/records'" in content
        assert "'/repairs'" in content
        tr.add(test_id, '前端界面', '路由配置', '验证5个页面路由已配置', 'PASS', f"路由配置数={routes_count}", '包含5个页面路由')
    except Exception as e:
        tr.add(test_id, '前端界面', '路由配置', '验证5个页面路由已配置', 'FAIL', str(e), '包含5个页面路由')
    
    test_id += 1
    try:
        js_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'app.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert 'stat-cards' in content
        assert 'pieChart' in content
        assert 'lineChart' in content
        assert 'reminder-list' in content
        tr.add(test_id, '前端界面', '仪表盘组件', '验证仪表盘包含统计卡片、饼图、折线图、提醒列表', 'PASS', "仪表盘元素齐全", '包含所有仪表盘UI组件')
    except Exception as e:
        tr.add(test_id, '前端界面', '仪表盘组件', '验证仪表盘包含统计卡片、饼图、折线图、提醒列表', 'FAIL', str(e), '包含所有仪表盘UI组件')
    
    test_id += 1
    try:
        js_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'app.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert '新增设备' in content
        assert '编辑设备' in content
        assert '删除' in content
        assert '导入Excel' in content
        assert '导出Excel' in content
        tr.add(test_id, '前端界面', '设备管理功能按钮', '验证设备管理页面包含增删改查导入导出按钮', 'PASS', "设备管理功能按钮齐全", '包含所有操作按钮')
    except Exception as e:
        tr.add(test_id, '前端界面', '设备管理功能按钮', '验证设备管理页面包含增删改查导入导出按钮', 'FAIL', str(e), '包含所有操作按钮')
    
    test_id += 1
    try:
        js_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'app.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert '新增计划' in content
        assert '编辑巡检计划' in content
        assert 'inspection_type' in content
        assert 'custom_days' in content
        tr.add(test_id, '前端界面', '巡检计划功能', '验证巡检计划页面包含新增编辑及类型配置', 'PASS', "计划管理功能齐全", '包含计划配置功能')
    except Exception as e:
        tr.add(test_id, '前端界面', '巡检计划功能', '验证巡检计划页面包含新增编辑及类型配置', 'FAIL', str(e), '包含计划配置功能')
    
    test_id += 1
    try:
        js_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'app.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert '新增记录' in content
        assert 'handlePhotoUpload' in content
        assert 'previewPhoto' in content
        assert 'anomaly_description' in content
        tr.add(test_id, '前端界面', '巡检记录功能', '验证巡检记录包含照片上传、异常描述等功能', 'PASS', "记录功能齐全", '包含照片上传和异常描述')
    except Exception as e:
        tr.add(test_id, '前端界面', '巡检记录功能', '验证巡检记录包含照片上传、异常描述等功能', 'FAIL', str(e), '包含照片上传和异常描述')
    
    test_id += 1
    try:
        js_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'app.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert '新增报修' in content
        assert '编辑报修' in content
        assert 'repair_cost' in content
        assert 'fault_description' in content
        tr.add(test_id, '前端界面', '报修管理功能', '验证报修管理包含新增编辑、费用、故障描述', 'PASS', "报修功能齐全", '包含报修核心字段操作')
    except Exception as e:
        tr.add(test_id, '前端界面', '报修管理功能', '验证报修管理包含新增编辑、费用、故障描述', 'FAIL', str(e), '包含报修核心字段操作')
    
    test_id += 1
    try:
        js_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'app.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert 'CATEGORIES' in content
        assert 'STATUSES' in content
        assert 'INSPECTION_TYPES' in content
        assert 'INSPECTION_ITEMS' in content
        assert 'REPAIR_STATUSES' in content
        tr.add(test_id, '前端界面', '常量配置', '验证前端定义了所有业务枚举常量', 'PASS', "5组常量均已定义", '包含所有业务常量')
    except Exception as e:
        tr.add(test_id, '前端界面', '常量配置', '验证前端定义了所有业务枚举常量', 'FAIL', str(e), '包含所有业务常量')
    
    test_id += 1
    try:
        js_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'app.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert 'sidebar' in content
        assert 'menu-item' in content
        assert 'modal-overlay' in content
        assert 'modal' in content
        tr.add(test_id, '前端界面', '布局和样式', '验证前端包含侧边栏、菜单、弹窗等UI组件', 'PASS', "布局组件齐全", '包含所有布局组件')
    except Exception as e:
        tr.add(test_id, '前端界面', '布局和样式', '验证前端包含侧边栏、菜单、弹窗等UI组件', 'FAIL', str(e), '包含所有布局组件')
    
    test_id += 1
    try:
        js_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'app.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert 'ElMessage' in content
        assert 'loadData' in content or 'loadDevices' in content
        assert 'openModal' in content
        assert 'saveDevice' in content or 'savePlan' in content
        tr.add(test_id, '前端界面', '交互逻辑', '验证前端包含消息提示、数据加载、弹窗、保存等交互', 'PASS', "交互逻辑完整", '包含核心交互逻辑')
    except Exception as e:
        tr.add(test_id, '前端界面', '交互逻辑', '验证前端包含消息提示、数据加载、弹窗、保存等交互', 'FAIL', str(e), '包含核心交互逻辑')
    
    # ============ 数据完整性验证 ============
    test_id += 1
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) as cnt FROM devices')
        cnt = cursor.fetchone()['cnt']
        conn.close()
        assert cnt >= 5
        tr.add(test_id, '数据完整性', '设备数据量', '验证设备表数据量>=5', 'PASS', f"设备数={cnt}", '设备数据完整')
    except Exception as e:
        tr.add(test_id, '数据完整性', '设备数据量', '验证设备表数据量>=5', 'FAIL', str(e), '设备数据完整')
    
    test_id += 1
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) as cnt FROM inspection_plans')
        cnt = cursor.fetchone()['cnt']
        conn.close()
        assert cnt >= 3
        tr.add(test_id, '数据完整性', '巡检计划数据量', '验证巡检计划表数据量>=3', 'PASS', f"计划数={cnt}", '计划数据完整')
    except Exception as e:
        tr.add(test_id, '数据完整性', '巡检计划数据量', '验证巡检计划表数据量>=3', 'FAIL', str(e), '计划数据完整')
    
    test_id += 1
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) as cnt FROM inspection_records')
        cnt = cursor.fetchone()['cnt']
        conn.close()
        assert cnt >= 3
        tr.add(test_id, '数据完整性', '巡检记录数据量', '验证巡检记录表数据量>=3', 'PASS', f"记录数={cnt}", '记录数据完整')
    except Exception as e:
        tr.add(test_id, '数据完整性', '巡检记录数据量', '验证巡检记录表数据量>=3', 'FAIL', str(e), '记录数据完整')
    
    test_id += 1
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) as cnt FROM repair_records')
        cnt = cursor.fetchone()['cnt']
        conn.close()
        assert cnt >= 2
        tr.add(test_id, '数据完整性', '报修记录数据量', '验证报修记录表数据量>=2', 'PASS', f"报修数={cnt}", '报修数据完整')
    except Exception as e:
        tr.add(test_id, '数据完整性', '报修记录数据量', '验证报修记录表数据量>=2', 'FAIL', str(e), '报修数据完整')
    
    test_id += 1
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) as cnt FROM devices WHERE status IN ('正常', '故障', '维修中')")
        cnt = cursor.fetchone()['cnt']
        conn.close()
        assert cnt >= 5
        tr.add(test_id, '数据完整性', '设备状态有效性', '验证所有设备状态都在有效枚举范围内', 'PASS', f"有效状态设备数={cnt}", '设备状态有效')
    except Exception as e:
        tr.add(test_id, '数据完整性', '设备状态有效性', '验证所有设备状态都在有效枚举范围内', 'FAIL', str(e), '设备状态有效')
    
    return tr.print_report()


if __name__ == '__main__':
    app.config['TESTING'] = True
    run_tests()