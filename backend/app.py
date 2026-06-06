from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import os
import json
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO

from database import get_db, init_db, calculate_next_date

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
EXPORT_FOLDER = os.path.join(os.path.dirname(__file__), 'exports')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORT_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def row_to_dict(row):
    return {key: row[key] for key in row.keys()}

@app.route('/')
def index():
    return jsonify({'message': '设备巡检管理系统 API'})

@app.route('/api/devices', methods=['GET'])
def get_devices():
    category = request.args.get('category')
    status = request.args.get('status')
    keyword = request.args.get('keyword')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = 'SELECT * FROM devices WHERE 1=1'
    params = []
    
    if category:
        query += ' AND category = ?'
        params.append(category)
    if status:
        query += ' AND status = ?'
        params.append(status)
    if keyword:
        query += ' AND (name LIKE ? OR asset_no LIKE ? OR brand LIKE ?)'
        params.extend([f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'])
    
    query += ' ORDER BY created_at DESC'
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    return jsonify([row_to_dict(row) for row in rows])

@app.route('/api/devices/<int:device_id>', methods=['GET'])
def get_device(device_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM devices WHERE id = ?', (device_id,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return jsonify({'error': '设备不存在'}), 404
    
    return jsonify(row_to_dict(row))

@app.route('/api/devices', methods=['POST'])
def create_device():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
        INSERT INTO devices (name, model, brand, asset_no, category, location, 
                            purchase_date, warranty_date, status, photo, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('name', ''),
            data.get('model', ''),
            data.get('brand', ''),
            data.get('asset_no', ''),
            data.get('category', '其他'),
            data.get('location', ''),
            data.get('purchase_date', ''),
            data.get('warranty_date', ''),
            data.get('status', '正常'),
            data.get('photo', ''),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ))
        conn.commit()
        device_id = cursor.lastrowid
        return jsonify({'id': device_id, 'message': '创建设备成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/devices/<int:device_id>', methods=['PUT'])
def update_device(device_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
        UPDATE devices SET name=?, model=?, brand=?, asset_no=?, category=?, 
                          location=?, purchase_date=?, warranty_date=?, status=?, 
                          photo=?, updated_at=? WHERE id=?
        ''', (
            data.get('name', ''),
            data.get('model', ''),
            data.get('brand', ''),
            data.get('asset_no', ''),
            data.get('category', '其他'),
            data.get('location', ''),
            data.get('purchase_date', ''),
            data.get('warranty_date', ''),
            data.get('status', '正常'),
            data.get('photo', ''),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            device_id
        ))
        conn.commit()
        return jsonify({'message': '更新设备成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/devices/<int:device_id>', methods=['DELETE'])
def delete_device(device_id):
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('DELETE FROM inspection_records WHERE device_id = ?', (device_id,))
        cursor.execute('DELETE FROM inspection_plans WHERE device_id = ?', (device_id,))
        cursor.execute('DELETE FROM repair_records WHERE device_id = ?', (device_id,))
        cursor.execute('DELETE FROM devices WHERE id = ?', (device_id,))
        conn.commit()
        return jsonify({'message': '删除设备成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/devices/import', methods=['POST'])
def import_devices():
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    if file and allowed_file(file.filename):
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            conn = get_db()
            cursor = conn.cursor()
            
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                if data.get('设备名称'):
                    cursor.execute('''
                    INSERT OR REPLACE INTO devices (name, model, brand, asset_no, category, 
                                                   location, purchase_date, warranty_date, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        str(data.get('设备名称', '')),
                        str(data.get('型号', '')),
                        str(data.get('品牌', '')),
                        str(data.get('资产编号', '')),
                        str(data.get('分类', '其他')),
                        str(data.get('位置', '')),
                        str(data.get('购买日期', '')),
                        str(data.get('保修期', '')),
                        str(data.get('状态', '正常'))
                    ))
                    count += 1
            
            conn.commit()
            conn.close()
            return jsonify({'message': f'成功导入 {count} 条记录'})
        except Exception as e:
            return jsonify({'error': str(e)}), 400
    
    return jsonify({'error': '不支持的文件类型'}), 400

@app.route('/api/plans', methods=['GET'])
def get_plans():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
    SELECT p.*, d.name as device_name, d.category as device_category 
    FROM inspection_plans p 
    LEFT JOIN devices d ON p.device_id = d.id 
    ORDER BY p.created_at DESC
    ''')
    rows = cursor.fetchall()
    conn.close()
    
    result = []
    for row in rows:
        d = row_to_dict(row)
        d['items'] = json.loads(d['items']) if d['items'] else []
        result.append(d)
    
    return jsonify(result)

@app.route('/api/plans/<int:plan_id>', methods=['GET'])
def get_plan(plan_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
    SELECT p.*, d.name as device_name 
    FROM inspection_plans p 
    LEFT JOIN devices d ON p.device_id = d.id 
    WHERE p.id = ?
    ''', (plan_id,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return jsonify({'error': '计划不存在'}), 404
    
    d = row_to_dict(row)
    d['items'] = json.loads(d['items']) if d['items'] else []
    return jsonify(d)

@app.route('/api/plans', methods=['POST'])
def create_plan():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    items_json = json.dumps(data.get('items', []), ensure_ascii=False)
    last_date = data.get('last_inspection_date') or datetime.now().strftime('%Y-%m-%d')
    next_date = calculate_next_date(last_date, data.get('inspection_type', '每周'), data.get('custom_days'))
    
    try:
        cursor.execute('''
        INSERT INTO inspection_plans (device_id, inspection_type, custom_days, items,
                                     last_inspection_date, next_inspection_date, responsible_person)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('device_id'),
            data.get('inspection_type', '每周'),
            data.get('custom_days'),
            items_json,
            data.get('last_inspection_date'),
            next_date,
            data.get('responsible_person', '')
        ))
        conn.commit()
        return jsonify({'id': cursor.lastrowid, 'message': '创建巡检计划成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/plans/<int:plan_id>', methods=['PUT'])
def update_plan(plan_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    items_json = json.dumps(data.get('items', []), ensure_ascii=False)
    last_date = data.get('last_inspection_date')
    if last_date:
        next_date = calculate_next_date(last_date, data.get('inspection_type', '每周'), data.get('custom_days'))
    else:
        next_date = None
    
    try:
        cursor.execute('''
        UPDATE inspection_plans SET device_id=?, inspection_type=?, custom_days=?, items=?,
                                   last_inspection_date=?, next_inspection_date=?, responsible_person=?
        WHERE id=?
        ''', (
            data.get('device_id'),
            data.get('inspection_type', '每周'),
            data.get('custom_days'),
            items_json,
            data.get('last_inspection_date'),
            next_date,
            data.get('responsible_person', ''),
            plan_id
        ))
        conn.commit()
        return jsonify({'message': '更新巡检计划成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/plans/<int:plan_id>', methods=['DELETE'])
def delete_plan(plan_id):
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('DELETE FROM inspection_plans WHERE id = ?', (plan_id,))
        conn.commit()
        return jsonify({'message': '删除巡检计划成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/records', methods=['GET'])
def get_records():
    device_id = request.args.get('device_id')
    result = request.args.get('result')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = '''
    SELECT r.*, d.name as device_name, d.category as device_category 
    FROM inspection_records r 
    LEFT JOIN devices d ON r.device_id = d.id 
    WHERE 1=1
    '''
    params = []
    
    if device_id:
        query += ' AND r.device_id = ?'
        params.append(device_id)
    if result:
        query += ' AND r.result = ?'
        params.append(result)
    
    query += ' ORDER BY r.inspection_date DESC, r.created_at DESC'
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    result_list = []
    for row in rows:
        d = row_to_dict(row)
        d['items_detail'] = json.loads(d['items_detail']) if d['items_detail'] else []
        d['photos'] = json.loads(d['photos']) if d['photos'] else []
        result_list.append(d)
    
    return jsonify(result_list)

@app.route('/api/records/<int:record_id>', methods=['GET'])
def get_record(record_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
    SELECT r.*, d.name as device_name 
    FROM inspection_records r 
    LEFT JOIN devices d ON r.device_id = d.id 
    WHERE r.id = ?
    ''', (record_id,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return jsonify({'error': '记录不存在'}), 404
    
    d = row_to_dict(row)
    d['items_detail'] = json.loads(d['items_detail']) if d['items_detail'] else []
    d['photos'] = json.loads(d['photos']) if d['photos'] else []
    return jsonify(d)

@app.route('/api/records', methods=['POST'])
def create_record():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    items_detail_json = json.dumps(data.get('items_detail', []), ensure_ascii=False)
    photos_json = json.dumps(data.get('photos', []), ensure_ascii=False)
    
    try:
        cursor.execute('''
        INSERT INTO inspection_records (device_id, plan_id, inspection_date, result,
                                       items_detail, anomaly_description, inspector, photos)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('device_id'),
            data.get('plan_id'),
            data.get('inspection_date', datetime.now().strftime('%Y-%m-%d')),
            data.get('result', '正常'),
            items_detail_json,
            data.get('anomaly_description', ''),
            data.get('inspector', ''),
            photos_json
        ))
        
        if data.get('result') == '异常':
            cursor.execute('UPDATE devices SET status = ? WHERE id = ?', 
                         ('故障', data.get('device_id')))
        
        if data.get('plan_id'):
            inspection_date = data.get('inspection_date', datetime.now().strftime('%Y-%m-%d'))
            cursor.execute('SELECT * FROM inspection_plans WHERE id = ?', (data.get('plan_id'),))
            plan = cursor.fetchone()
            if plan:
                next_date = calculate_next_date(inspection_date, plan['inspection_type'], plan['custom_days'])
                cursor.execute('''
                UPDATE inspection_plans 
                SET last_inspection_date = ?, next_inspection_date = ?
                WHERE id = ?
                ''', (inspection_date, next_date, data.get('plan_id')))
        
        conn.commit()
        return jsonify({'id': cursor.lastrowid, 'message': '创建巡检记录成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/records/<int:record_id>', methods=['DELETE'])
def delete_record(record_id):
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('DELETE FROM inspection_records WHERE id = ?', (record_id,))
        conn.commit()
        return jsonify({'message': '删除巡检记录成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/repairs', methods=['GET'])
def get_repairs():
    device_id = request.args.get('device_id')
    status = request.args.get('status')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = '''
    SELECT r.*, d.name as device_name, d.category as device_category 
    FROM repair_records r 
    LEFT JOIN devices d ON r.device_id = d.id 
    WHERE 1=1
    '''
    params = []
    
    if device_id:
        query += ' AND r.device_id = ?'
        params.append(device_id)
    if status:
        query += ' AND r.status = ?'
        params.append(status)
    
    query += ' ORDER BY r.report_date DESC, r.created_at DESC'
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    return jsonify([row_to_dict(row) for row in rows])

@app.route('/api/repairs/<int:repair_id>', methods=['GET'])
def get_repair(repair_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
    SELECT r.*, d.name as device_name 
    FROM repair_records r 
    LEFT JOIN devices d ON r.device_id = d.id 
    WHERE r.id = ?
    ''', (repair_id,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return jsonify({'error': '记录不存在'}), 404
    
    return jsonify(row_to_dict(row))

@app.route('/api/repairs', methods=['POST'])
def create_repair():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
        INSERT INTO repair_records (device_id, report_date, fault_description, 
                                   repair_date, repair_result, repair_cost, status)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('device_id'),
            data.get('report_date', datetime.now().strftime('%Y-%m-%d')),
            data.get('fault_description', ''),
            data.get('repair_date'),
            data.get('repair_result', ''),
            data.get('repair_cost', 0),
            data.get('status', '待处理')
        ))
        
        status = data.get('status', '待处理')
        device_status = '维修中' if status in ['待处理', '处理中'] else '正常'
        cursor.execute('UPDATE devices SET status = ? WHERE id = ?', 
                     (device_status, data.get('device_id')))
        
        conn.commit()
        return jsonify({'id': cursor.lastrowid, 'message': '创建报修记录成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/repairs/<int:repair_id>', methods=['PUT'])
def update_repair(repair_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
        UPDATE repair_records SET device_id=?, report_date=?, fault_description=?,
                                 repair_date=?, repair_result=?, repair_cost=?, status=?
        WHERE id=?
        ''', (
            data.get('device_id'),
            data.get('report_date'),
            data.get('fault_description'),
            data.get('repair_date'),
            data.get('repair_result'),
            data.get('repair_cost', 0),
            data.get('status', '待处理'),
            repair_id
        ))
        
        status = data.get('status', '待处理')
        device_status = '维修中' if status in ['待处理', '处理中'] else '正常'
        cursor.execute('UPDATE devices SET status = ? WHERE id = ?', 
                     (device_status, data.get('device_id')))
        
        conn.commit()
        return jsonify({'message': '更新报修记录成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/repairs/<int:repair_id>', methods=['DELETE'])
def delete_repair(repair_id):
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('SELECT device_id FROM repair_records WHERE id = ?', (repair_id,))
        row = cursor.fetchone()
        if row:
            cursor.execute('DELETE FROM repair_records WHERE id = ?', (repair_id,))
        conn.commit()
        return jsonify({'message': '删除报修记录成功'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/dashboard/stats', methods=['GET'])
def get_dashboard_stats():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT COUNT(*) as total FROM devices')
    total_devices = cursor.fetchone()['total']
    
    cursor.execute("SELECT COUNT(*) as normal FROM devices WHERE status = '正常'")
    normal_devices = cursor.fetchone()['normal']
    
    cursor.execute("SELECT COUNT(*) as anomaly FROM devices WHERE status IN ('故障', '维修中')")
    anomaly_devices = cursor.fetchone()['anomaly']
    
    today = datetime.now().strftime('%Y-%m-%d')
    cursor.execute('''
    SELECT COUNT(*) as today_count 
    FROM inspection_plans 
    WHERE next_inspection_date = ?
    ''', (today,))
    today_inspection = cursor.fetchone()['today_count']
    
    cursor.execute('''
    SELECT COUNT(*) as overdue 
    FROM inspection_plans 
    WHERE next_inspection_date < ?
    ''', (today,))
    overdue_count = cursor.fetchone()['overdue']
    
    cursor.execute('''
    SELECT category, COUNT(*) as count 
    FROM devices 
    GROUP BY category
    ''')
    category_stats = [row_to_dict(row) for row in cursor.fetchall()]
    
    cursor.execute('''
    SELECT strftime('%Y-%m-%d', inspection_date) as date, 
           COUNT(*) as total,
           SUM(CASE WHEN result = '正常' THEN 1 ELSE 0 END) as completed
    FROM inspection_records
    WHERE inspection_date >= date('now', '-30 days')
    GROUP BY date
    ORDER BY date
    ''')
    trend_data = [row_to_dict(row) for row in cursor.fetchall()]
    
    cursor.execute('''
    SELECT COUNT(*) as warranty_soon 
    FROM devices 
    WHERE warranty_date != '' 
    AND warranty_date IS NOT NULL 
    AND date(warranty_date) BETWEEN date('now') AND date('now', '+30 days')
    ''')
    warranty_soon = cursor.fetchone()['warranty_soon']
    
    conn.close()
    
    return jsonify({
        'total_devices': total_devices,
        'normal_devices': normal_devices,
        'anomaly_devices': anomaly_devices,
        'today_inspection': today_inspection,
        'overdue_count': overdue_count,
        'category_stats': category_stats,
        'trend_data': trend_data,
        'warranty_soon': warranty_soon
    })

@app.route('/api/dashboard/reminders', methods=['GET'])
def get_reminders():
    conn = get_db()
    cursor = conn.cursor()
    today = datetime.now().strftime('%Y-%m-%d')
    
    cursor.execute('''
    SELECT p.*, d.name as device_name 
    FROM inspection_plans p 
    LEFT JOIN devices d ON p.device_id = d.id 
    WHERE p.next_inspection_date <= date('now', '+7 days')
    ORDER BY p.next_inspection_date
    ''')
    inspection_reminders = [row_to_dict(row) for row in cursor.fetchall()]
    
    cursor.execute('''
    SELECT * FROM devices 
    WHERE warranty_date != '' 
    AND warranty_date IS NOT NULL 
    AND date(warranty_date) BETWEEN date('now') AND date('now', '+30 days')
    ORDER BY warranty_date
    ''')
    warranty_reminders = [row_to_dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    return jsonify({
        'inspection_reminders': inspection_reminders,
        'warranty_reminders': warranty_reminders
    })

@app.route('/api/export/devices', methods=['GET'])
def export_devices():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM devices ORDER BY created_at DESC')
    rows = cursor.fetchall()
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '设备清单'
    
    headers = ['ID', '设备名称', '型号', '品牌', '资产编号', '分类', '位置', '购买日期', '保修期', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row['id'])
        ws.cell(row=row_idx, column=2, value=row['name'])
        ws.cell(row=row_idx, column=3, value=row['model'])
        ws.cell(row=row_idx, column=4, value=row['brand'])
        ws.cell(row=row_idx, column=5, value=row['asset_no'])
        ws.cell(row=row_idx, column=6, value=row['category'])
        ws.cell(row=row_idx, column=7, value=row['location'])
        ws.cell(row=row_idx, column=8, value=row['purchase_date'])
        ws.cell(row=row_idx, column=9, value=row['warranty_date'])
        ws.cell(row=row_idx, column=10, value=row['status'])
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    filename = f'devices_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(EXPORT_FOLDER, filename)
    wb.save(filepath)
    
    return send_from_directory(EXPORT_FOLDER, filename, as_attachment=True)

@app.route('/api/export/records', methods=['GET'])
def export_records():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
    SELECT r.*, d.name as device_name 
    FROM inspection_records r 
    LEFT JOIN devices d ON r.device_id = d.id 
    ORDER BY r.inspection_date DESC
    ''')
    rows = cursor.fetchall()
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '巡检记录'
    
    headers = ['ID', '设备名称', '巡检日期', '巡检结果', '异常描述', '巡检人']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row['id'])
        ws.cell(row=row_idx, column=2, value=row['device_name'])
        ws.cell(row=row_idx, column=3, value=row['inspection_date'])
        ws.cell(row=row_idx, column=4, value=row['result'])
        ws.cell(row=row_idx, column=5, value=row['anomaly_description'])
        ws.cell(row=row_idx, column=6, value=row['inspector'])
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    filename = f'records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(EXPORT_FOLDER, filename)
    wb.save(filepath)
    
    return send_from_directory(EXPORT_FOLDER, filename, as_attachment=True)

@app.route('/api/export/repairs', methods=['GET'])
def export_repairs():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
    SELECT r.*, d.name as device_name 
    FROM repair_records r 
    LEFT JOIN devices d ON r.device_id = d.id 
    ORDER BY r.report_date DESC
    ''')
    rows = cursor.fetchall()
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '报修记录'
    
    headers = ['ID', '设备名称', '报修日期', '故障描述', '维修日期', '维修结果', '维修费用', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row['id'])
        ws.cell(row=row_idx, column=2, value=row['device_name'])
        ws.cell(row=row_idx, column=3, value=row['report_date'])
        ws.cell(row=row_idx, column=4, value=row['fault_description'])
        ws.cell(row=row_idx, column=5, value=row['repair_date'])
        ws.cell(row=row_idx, column=6, value=row['repair_result'])
        ws.cell(row=row_idx, column=7, value=row['repair_cost'])
        ws.cell(row=row_idx, column=8, value=row['status'])
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    filename = f'repairs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(EXPORT_FOLDER, filename)
    wb.save(filepath)
    
    return send_from_directory(EXPORT_FOLDER, filename, as_attachment=True)

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        return jsonify({'url': f'/uploads/{filename}', 'filename': filename})
    
    return jsonify({'error': '不支持的文件类型'}), 400

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=8000)
